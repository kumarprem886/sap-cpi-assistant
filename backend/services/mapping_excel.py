"""
Generate a mapping specification Excel (.xlsx) from a CPI iFlow ZIP.
Format mirrors SAP CPI's built-in "Export as Excel" button:
  - The .mmap XML inside the ZIP is the exact same source SAP uses
  - Supports direct field mappings AND nested function call expressions
"""
import io, re, zipfile


# ── mmap XML parser ────────────────────────────────────────────────────────────

def _find_close(xml: str, pos: int) -> int:
    """
    Starting from pos (the character right after a '<brick' opener),
    find the index of the matching '</brick>' by counting nesting depth.
    """
    depth = 1
    i = pos
    while i < len(xml) and depth > 0:
        o = xml.find('<brick', i)
        c = xml.find('</brick>', i)
        if c == -1:
            return -1
        if o != -1 and o < c:
            # Another <brick starts before this close
            depth += 1
            i = o + 6          # skip past '<brick'
        else:
            depth -= 1
            if depth == 0:
                return c
            i = c + 8          # skip past '</brick>'
    return -1


def _build_expr(xml_frag: str) -> str:
    """
    Recursively build a mapping expression string from an XML fragment
    that represents one brick (Src / Dst / Std / Const).

    This replicates the output of SAP CPI's "Export as Excel" format:
      - Direct field:  /source/field/path
      - Function:      funcName(\narg1,\narg2)
      - Constant:      "value"
    """
    xml_frag = xml_frag.strip()
    if not xml_frag:
        return ''

    # Read the opening <brick ...> tag
    tag_m = re.match(r'<brick\b([^>]*)>', xml_frag)
    if not tag_m:
        # Might be a self-closing <brick .../>
        sc_m = re.match(r'<brick\b([^>]*)/>', xml_frag)
        if sc_m:
            attrs = sc_m.group(1)
            path_m = re.search(r'\bpath="([^"]+)"', attrs)
            return path_m.group(1) if path_m else ''
        return ''

    attrs = tag_m.group(1)
    btype = re.search(r'\btype="([^"]+)"', attrs)
    btype = btype.group(1) if btype else ''
    bname = re.search(r'\bname="([^"]+)"', attrs)
    bname = bname.group(1) if bname else ''
    bpath = re.search(r'\bpath="([^"]+)"', attrs)
    bpath = bpath.group(1) if bpath else ''

    tag_end = tag_m.end()
    # Get the inner content of this brick
    close_pos = _find_close(xml_frag, tag_end)
    if close_pos == -1:
        inner = xml_frag[tag_end:]
    else:
        inner = xml_frag[tag_end:close_pos]

    if btype == 'Src':
        return bpath

    if btype in ('Dst', ''):
        # Return the expression of its <arg> content — greedy to avoid inner </arg> cut-off
        arg_m = re.search(r'<arg>(.*)</arg>', inner, re.DOTALL)
        if arg_m is None:
            return bpath or ''
        return _build_arg_exprs(arg_m.group(1))

    if btype == 'Std':
        # Standard function: funcName(\narg1,\narg2)
        # Use greedy .* so nested <arg>...</arg> pairs are fully captured
        arg_m = re.search(r'<arg>(.*)</arg>', inner, re.DOTALL)
        if arg_m is None:
            return f'{bname}()'
        args = _split_bricks(arg_m.group(1))  # now also returns <const> fragments
        arg_exprs = []
        for frag in args:
            frag = frag.strip()
            if frag.startswith('<brick'):
                arg_exprs.append(_build_expr(frag))
            elif frag.startswith('<const'):
                cv = re.search(r'<const[^>]*>([^<]*)</const>', frag)
                if cv:
                    # Unescape HTML entities in constant values
                    val = cv.group(1).replace('&quot;', '"').replace('&amp;', '&')
                    arg_exprs.append(f'"{val}"')
        if not arg_exprs:
            return f'{bname}()'
        return f'{bname}(\n' + ',\n'.join(arg_exprs) + ')'

    return bpath or bname or ''


def _build_arg_exprs(arg_inner: str) -> str:
    """Build expression from the content of an <arg> element."""
    frags = _split_bricks(arg_inner)
    exprs = [_build_expr(f.strip()) for f in frags if f.strip()]
    if len(exprs) == 1:
        return exprs[0]
    return ',\n'.join(exprs)


def _split_bricks(xml: str) -> list:
    """
    Split an XML fragment into its top-level children: <brick> and <const>.
    Uses depth counting for bricks to handle nesting correctly.
    """
    results = []
    i = 0
    while i < len(xml):
        b_start = xml.find('<brick', i)
        c_start = xml.find('<const', i)

        # Pick whichever comes first
        if b_start == -1 and c_start == -1:
            break
        if b_start == -1:
            next_start, is_const = c_start, True
        elif c_start == -1:
            next_start, is_const = b_start, False
        else:
            next_start, is_const = (c_start, True) if c_start < b_start else (b_start, False)

        if is_const:
            c_end = xml.find('</const>', next_start)
            if c_end == -1:
                results.append(xml[next_start:])
                break
            results.append(xml[next_start: c_end + 8])
            i = c_end + 8
            continue

        # It's a <brick ...>
        tag_end_m = re.search(r'>', xml[next_start:])
        if not tag_end_m:
            break
        tag_end = next_start + tag_end_m.end()

        # Self-closing?
        if xml[tag_end - 2] == '/':
            results.append(xml[next_start:tag_end])
            i = tag_end
            continue

        close = _find_close(xml, tag_end)
        if close == -1:
            results.append(xml[next_start:])
            break
        results.append(xml[next_start: close + 8])
        i = close + 8
    return results


def _parse_mmap_xml(mmap_xml: str) -> list[tuple[str, str]]:
    """
    Parse the .mmap XML and return (target_path, mapping_expression) pairs.

    Supports:
      • Direct 1-to-1 field mapping  →  source_field_path
      • Standard function calls       →  funcName(\narg1,\narg2)
      • Nested function expressions   →  outer(\ninner(\nfield))
      • Constants                     →  "value"

    This replicates exactly what SAP CPI's "Export as Excel" button produces.
    """
    pairs: list[tuple[str, str]] = []

    # Extract transformation section
    trans_m = re.search(r'<transformation>(.*?)</transformation>', mmap_xml, re.DOTALL)
    if not trans_m:
        return pairs

    trans = trans_m.group(1)

    # Iterate over top-level Dst bricks
    for frag in _split_bricks(trans):
        frag = frag.strip()
        if not frag:
            continue
        tag_m = re.match(r'<brick\b([^>]*)>', frag)
        if not tag_m:
            continue
        attrs = tag_m.group(1)
        btype = re.search(r'\btype="([^"]+)"', attrs)
        bpath = re.search(r'\bpath="([^"]+)"', attrs)
        if not btype or btype.group(1) != 'Dst' or not bpath:
            continue

        dst_path = bpath.group(1)

        # Get inner content
        tag_end = tag_m.end()
        close = _find_close(frag, tag_end)
        inner = frag[tag_end:close] if close != -1 else frag[tag_end:]

        # Get first <arg> element — use greedy so nested </arg> don't terminate early
        arg_m = re.search(r'<arg>(.*)</arg>', inner, re.DOTALL)
        if arg_m is None:
            pairs.append((dst_path, ''))
            continue

        expr = _build_arg_exprs(arg_m.group(1))
        pairs.append((dst_path, expr))

    return pairs


# ── Excel generation ───────────────────────────────────────────────────────────

def _extract_mmap_meta(mmap_xml: str) -> dict:
    """
    Extract metadata from the mmap XML:
      - source_msg:  root element name of the source message
      - target_msg:  root element name of the target message
      - source_file: XSD filename for source
      - target_file: XSD filename for target
      - description: mapping description (if any)
    """
    meta = {
        'source_msg': '', 'target_msg': '',
        'source_file': '', 'target_file': '',
        'description': '',
    }

    # Description
    desc_m = re.search(r'<description>([^<]*)</description>', mmap_xml)
    if desc_m:
        meta['description'] = desc_m.group(1).strip()

    # lnkRole sections  (SOURCE_IFR_MESS / TARGET_IFR_MESS)
    for role_m in re.finditer(
        r'<lnkRole[^>]*\brole="(SOURCE_IFR_MESS|TARGET_IFR_MESS)"[^>]*>(.*?)</lnkRole>',
        mmap_xml, re.DOTALL
    ):
        role   = role_m.group(1)
        body   = role_m.group(2)
        # Extract <elem> values: [filename, path, root-element]
        elems  = re.findall(r'<elem>([^<]*)</elem>', body)
        if not elems:
            continue
        xsd_file = elems[0] if len(elems) > 0 else ''
        root_msg = elems[2] if len(elems) > 2 else elems[0]

        if role == 'SOURCE_IFR_MESS':
            meta['source_file'] = xsd_file
            meta['source_msg']  = root_msg
        else:
            meta['target_file'] = xsd_file
            meta['target_msg']  = root_msg

    return meta


def generate_mapping_excel(iflow_zip_bytes: bytes) -> tuple[bytes, str] | None:
    """
    Parse the iFlow ZIP, find the .mmap file, and generate an Excel mapping spec.
    Returns (xlsx_bytes, filename) or None if no .mmap found.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    with zipfile.ZipFile(io.BytesIO(iflow_zip_bytes)) as z:
        files = z.namelist()
        mmap_files = [f for f in files if f.endswith('.mmap')]
        if not mmap_files:
            return None

        mmap_path = mmap_files[0]
        mmap_name = mmap_path.split('/')[-1].replace('.mmap', '')
        mmap_xml  = z.read(mmap_path).decode('utf-8', 'replace')

    # Extract full field pair list — filter to only mapped rows (non-empty source)
    all_pairs = _parse_mmap_xml(mmap_xml)
    pairs = [(dst, src) for dst, src in all_pairs if src.strip()]

    meta = _extract_mmap_meta(mmap_xml)

    # ── Workbook ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview"

    SAP_DARK  = "1A1A2E"
    SAP_BLUE  = "006DB3"
    COL_FILL  = PatternFill("solid", fgColor="2C3E6B")
    EVEN_FILL = PatternFill("solid", fgColor="EEF2F7")
    ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
    HDR_FILL  = PatternFill("solid", fgColor=SAP_DARK)
    SECT_FILL = PatternFill("solid", fgColor=SAP_BLUE)
    WHITE_FT  = Font(color="FFFFFF", bold=True,  name="Calibri", size=10)
    BOLD_FT   = Font(bold=True,  name="Calibri", size=10)
    NORM_FT   = Font(name="Calibri", size=9)
    MONO_FT   = Font(name="Courier New", size=9)   # for function expressions
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    LEFT_WRAP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(row, col, val, fill=HDR_FILL, ft=WHITE_FT, al=LEFT):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = ft; c.alignment = al; c.border = BORDER

    def _cell(row, col, val, fill=ODD_FILL, ft=NORM_FT, al=LEFT):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = ft; c.alignment = al; c.border = BORDER

    # Row 1: Title
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="OVERVIEW")
    c.fill = HDR_FILL
    c.font = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    c.alignment = CENTER; c.border = BORDER

    # Row 2: Column headers
    _hdr(2, 1, "PARAMETER", COL_FILL); _hdr(2, 2, "VALUE", COL_FILL); _hdr(2, 3, "", COL_FILL)

    # Rows 3-8: Metadata (using values extracted from mmap lnkRole sections)
    meta_rows = [
        ("Name of Mapping",            f"{mmap_name}.mmap"),
        ("Description",                meta['description']),
        ("Name of the Source Messages", f"[{meta['source_msg']}]" if meta['source_msg'] else ""),
        ("Name of the Source Files",    f"[{meta['source_file']}]" if meta['source_file'] else ""),
        ("Name of the Target Messages", f"[{meta['target_msg']}]" if meta['target_msg'] else ""),
        ("Name of the Target Files",    f"[{meta['target_file']}]" if meta['target_file'] else ""),
    ]
    meta = meta_rows
    for i, (label, value) in enumerate(meta):
        row  = 3 + i
        fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
        _cell(row, 1, label, fill=fill, ft=BOLD_FT)
        _cell(row, 2, value, fill=fill)
        _cell(row, 3, "",    fill=fill)

    ws.row_dimensions[9].height = 6

    # Row 10: DEFINITION
    ws.merge_cells("A10:C10")
    c = ws.cell(row=10, column=1, value="DEFINITION")
    c.fill = SECT_FILL; c.font = WHITE_FT; c.alignment = CENTER; c.border = BORDER

    # Row 11: Column headers
    _hdr(11, 1, "TARGET",           COL_FILL, al=CENTER)
    _hdr(11, 2, "MAPPING (SOURCE)", COL_FILL, al=CENTER)
    _hdr(11, 3, "TYPE",             COL_FILL, al=CENTER)

    # Rows 12+: Mapping pairs
    for i, (dst, src) in enumerate(pairs):
        row  = 12 + i
        fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
        _cell(row, 1, dst, fill=fill)
        # Use monospace font for function expressions; detect by newline presence
        is_func = '\n' in src
        c2 = ws.cell(row=row, column=2, value=src)
        c2.fill = fill
        c2.font = MONO_FT if is_func else NORM_FT
        c2.alignment = LEFT_WRAP if is_func else LEFT
        c2.border = BORDER
        _cell(row, 3, "", fill=fill)

    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12

    # Row heights
    ws.row_dimensions[1].height  = 24
    ws.row_dimensions[2].height  = 16
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 16

    ws.freeze_panes = "A12"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"{mmap_name}_MappingSpec.xlsx"
